# _*_ coding.utf-8 _*_
# Editor: Ziqi LIU
# Date: 2020/6/2
# File name: StepOne.py
# Tool: PyCharm

import openpyxl
from openpyxl.utils import column_index_from_string

wb = openpyxl.load_workbook('D:\pycharm\Question1Material.xlsx')  # 打开工作表生成工作簿对象
rawSheet = wb.active

for row in range(2, rawSheet.max_row + 1):  # 越过标题行，遍历数据行
    rowS = str(row)

    # 计算违约指数
    rawSheet.cell(row=row, column=column_index_from_string('H'), value='=E' + rowS + '+F' + rowS)
    # 单支债券新预警时间
    rawSheet.cell(row=row, column=column_index_from_string('I'),
                  value='=IF(H' + rowS + '>=50, A' + rowS + ', "")')

wb.save('D:\pycharm\Deloitte\Step1.xlsx')